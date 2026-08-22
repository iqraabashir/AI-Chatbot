from openpyxl import load_workbook
import sqlite3

DATABASE = "chatbot/knowledge.db"

conn = sqlite3.connect(DATABASE)
cursor = conn.cursor()

workbook = load_workbook(
    "data/official_data/knowledge/admissions.xlsx"
)

sheet = workbook.active

count = 0

for row in sheet.iter_rows(min_row=2, values_only=True):

    if row[0] is None:
        continue

    cursor.execute("""

    INSERT INTO admissions(

        category,
        topic,
        applicable_to,
        description,
        required_documents,
        eligibility_basis,
        admission_mode,
        selection_basis,
        reservation,
        application_fee,
        admission_portal,
        important_notes,
        source,
        url,
        last_updated,
        keywords,
        question_examples,
        programme_level

    )
    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)

    """, row)

    count += 1

conn.commit()
conn.close()

print(f"{count} Admission Records Imported Successfully.")