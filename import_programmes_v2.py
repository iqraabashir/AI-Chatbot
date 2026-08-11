from openpyxl import load_workbook
import sqlite3

DATABASE = "chatbot/knowledge.db"

conn = sqlite3.connect(DATABASE)
cursor = conn.cursor()

workbook = load_workbook(
    "data/official_data/knowledge/programme_master.xlsx"
)

sheet = workbook.active

count = 0

for row in sheet.iter_rows(min_row=2, values_only=True):

    if row[0] is None:
        continue

    cursor.execute("""
    INSERT INTO academic_programmes
   (
     programme,
     specialization,
     level,
     college,
     department,
     duration,
     programme_type,
     school,
     campus,
     eligibility,
     intake,
     fee,
     admission_process,
     selection_process,
     overview,
     subject_overview,
     source,
     url,
     last_updated
    )

    VALUES
     (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)

    """, row)

    count += 1

conn.commit()
conn.close()

print(f"{count} Programmes Imported Successfully.")