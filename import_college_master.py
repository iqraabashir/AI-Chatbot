from openpyxl import load_workbook
import sqlite3

DATABASE = "chatbot/knowledge.db"

conn = sqlite3.connect(DATABASE)
cursor = conn.cursor()

workbook = load_workbook(
    "data/official_data/knowledge/colleges.xlsx"
)

sheet = workbook.active

count = 0

for row in sheet.iter_rows(min_row=2, values_only=True):

    if row[0] is None:
        continue

    cursor.execute("""

    INSERT INTO college_master(

        college_name,
        short_name,
        university,
        established,
        type,
        campus,
        address,
        district,
        state,
        principal,
        overview,
        departments,
        programmes_offered,
        library,
        hostel,
        laboratories,
        sports,
        ncc,
        nss,
        facilities,
        website,
        google_maps,
        official_email

    )
    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, row)
    count += 1
conn.commit()
conn.close()
print(f"{count} Colleges Imported Successfully.")