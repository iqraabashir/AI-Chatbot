from openpyxl import load_workbook
from chatbot.knowledge_database import add_topic, add_field, add_value

workbook = load_workbook("data/official_data/knowledge/library.xlsx")
sheet = workbook.active


topics = {}

for row in sheet.iter_rows(min_row=2, values_only=True):

    topic_name, field_name, field_value, source, url, last_updated = row

    if topic_name not in topics:
        topic_id = add_topic(
            topic_name,
            "Library",
            f"{topic_name} information"
        )
        topics[topic_name] = topic_id

    topic_id = topics[topic_name]

    field_id = add_field(topic_id, field_name)

    add_value(
        field_id,
        field_value,
        source,
        url,
        str(last_updated)
    )

print("Library Imported Successfully!")