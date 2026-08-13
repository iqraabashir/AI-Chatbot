from openpyxl import load_workbook
import sqlite3

DATABASE = "chatbot/knowledge.db"

EXCEL_FILE = "data/official_data/knowledge/university_info.xlsx"

conn = sqlite3.connect(DATABASE)
cursor = conn.cursor()

workbook = load_workbook(EXCEL_FILE)
sheet = workbook.active

count = 0

for row in sheet.iter_rows(min_row=2, values_only=True):

    if row[0] is None:
        continue

    intent = row[0]
    topic = row[1]
    field = row[2]
    value = row[3]
    source = row[4]
    url = row[5]
    email = row[6]

    cursor.execute("""
        INSERT INTO university_info (
            intent,
            topic,
            field,
            value,
            source,
            url,
            email
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        intent,
        topic,
        field,
        value,
        source,
        url,
        email
    ))

    count += 1

conn.commit()
conn.close()

print(f"{count} University Information Records Imported Successfully.")